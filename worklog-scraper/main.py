import openpyxl
import Demo
import calendar
import glob2
import time
import os
from datetime import datetime
from datetime import date

month = 3
year = 2022
machine = "Laminator".upper()
monthSel = calendar.month_name[month]
genDir = f"{year}_{monthSel}/"


def loggera(name, listOfEQ={}, details={}):
    book_names = glob2.glob(f'{genDir}/*.xlsx')
    book_names.sort()
    alt = {"E1-US-LA-01": "E1 Lami1", "E1-US-LA-02": "E1 Lami2",
           "E1-US-LA-03": "E1 Lami3", "E1-US-LA-04": "E1 Lami4",
           "E2-US-LA-01": "E2 Lami1", "E2-US-LA-02": "E2 Lami2",
           "E2-US-LA-03": "E2 Lami3", "E2-US-LA-04": "E2 Lami4",
           "E3-US-LA-01": "E3 Lami1", "E3-US-LA-02": "E3 Lami2",
           "E3-US-LA-03": "E3 Lami3", "E3-US-LA-04": "E3 Lami4"
           }
    # Book names is the directory for every worklog in a given month. We will loop through it and save data to objects
    # "ListofEQ" and "details"
    machine_col, start_time_col, end_time_col,repair_time_col,equipment_name_col = None,None,None,None,None
    root_cause_col, issue_col, issue_details_col = None,None,None

    for index, nm in enumerate(book_names):

        wb = openpyxl.load_workbook(nm)
        ws = wb[wb.sheetnames[0]]

        # After setting this the first time, we dont need to do it again

        if not machine_col:
            i = 1
            while ws.cell(row=7, column=i).value is not None:
                end_time_dic = {}
                # Loop through and find Machine name, start time, and end time
                if ws.cell(row=7, column=i).value == "Machine":
                    machine_col = i
                if ws.cell(row=7, column=i).value == "StartTime":
                    start_time_col = i
                if ws.cell(row=7, column=i).value == "EndTime":
                    end_time_col = i
                if ws.cell(row=7, column=i).value == "Repair Time (minutes)":
                    repair_time_col = i
                if ws.cell(7, i).value == "EQ":
                    equipment_name_col = i
                if ws.cell(7, i).value == "Issues Details":
                    issue_details_col = i
                if ws.cell(7, i).value == "Root Causes":
                    root_cause_col = i
                if ws.cell(7, i).value == "Issues":
                    issue_col = i
                i += 1
        # Not I want to go down the rows
        i = 8
        # row, column
        temp = dict()
        mdeets = dict()
        while ws.cell(i, 1).value is not None:

            start_time, end_time = None, None
            # As we go down the rows, if we find something in the row at the machine col that matches, do the method V
            if ws.cell(i, machine_col).value == machine:
                machine_name = alt[ws.cell(i, equipment_name_col).value]
                # Get the starting and ending time
                start_time = ws.cell(i, start_time_col).value
                end_time = ws.cell(i, end_time_col).value
                if machine_name in temp:
                    # If this specific equipment is already in temp, add the minutes logged up
                    temp[machine_name] += int(ws.cell(i, repair_time_col).value)
                if machine_name not in temp:
                    # If this specific equipment is not already in the object, create it and set the initial downtime
                    temp[machine_name] = int(ws.cell(i, repair_time_col).value)

                if end_time.split(":")[1] == "59":
                    # If the time is x:59, change the end time to x+1:00.
                    end_time = ":".join([str((int(end_time.split(":")[0]) + 1) % 24), "00"])
                    if int(end_time.split(":")[0]) < 10:
                        end_time = f"0{end_time}"
                # Add the end time to the end time dictionary
                try:
                    end_time_dic[ws.cell(i, equipment_name_col).value].append(end_time)
                    # print(end_time)
                except:
                    end_time_dic[ws.cell(i, equipment_name_col).value] = [end_time]
                if start_time in end_time_dic[ws.cell(i, equipment_name_col).value]:
                    # If the start time is the same time as the end time of another interval with the same machine, skip
                    # Clean up
                    #print("TEST >", machine_name, start_time, end_time)
                    i += 1
                    continue
                #print("=============next+++++++++")
                ########################################
                # If the start time is not the same as the end time of the same machine (noncontinuous) create a new log
                # -------- NEW LOG -------- #

                # For details, I need to structure it in a way that doesnt break everything
                # format: Start time -> Root cause, Issue: Issue Details
                if machine_name in mdeets:
                    mdeets[machine_name].append(
                        f"{start_time}->{ws.cell(i, root_cause_col).value},{ws.cell(i, issue_col).value}: "
                        f"{ws.cell(i, issue_details_col).value}")

                if machine_name not in mdeets:
                    mdeets[machine_name] = [
                        f"{start_time}->{ws.cell(i, root_cause_col).value},{ws.cell(i, issue_col).value}: "
                        f"{ws.cell(i, issue_details_col).value}"]
            i += 1

        # Loop through the file. As long as there is still data in that row, execute the following lines

        # List of equipment that has broken down
        listOfEQ[index + 1] = temp
        # Detailed description of what happened and what was done to fix it
        details[index + 1] = mdeets
    return listOfEQ, details


def deleteDirectoryFiles():
    # Continue by deleting all files in Sorted directory
    Directory = glob2.glob(genDir + "*")
    for i in Directory:
        os.remove(i)


def deletewkfiles():
    Directoryb = glob2.glob('Worklog_Files\*')
    for i in Directoryb:
        os.remove(i)


def WorkBookCompiler():
    wBookComp = openpyxl.Workbook()
    time.sleep(1)
    bookNames = glob2.glob('Worklog_Files/*.xlsx')
    '''Now that i have a list of all the names, I need to go through each and
    Add them to the master workbook "wBookComp" Then rename the to their pages to
    year month and day'''

    i = 0

    for a in bookNames:
        i = 0
        roe = 8
        tempbook = openpyxl.load_workbook(a)
        tempsheet = tempbook["Equipment Hourly Worklog Report"]
        datesheet = str(tempsheet['A4'].value)[16: 24]
        tempbook.save(genDir + datesheet + ".xlsx")
    sheetnames = glob2.glob(genDir + '/*.xlsx')
    return [len(sheetnames), sheetnames]


# Grab list of days given month and year up to current day or end of month
# This then logs into the website and downloads the days on worklog


'''Grab info from work files directory and put in sorted'''


def main():
    invdictionary2 = {"E1 Lami1": 4, "E1 Lami2": 5, "E1 Lami3": 6, "E1 Lami4": 7,
                      "E2 Lami1": 8, "E2 Lami2": 9, "E2 Lami3": 10, "E2 Lami4": 11,
                      "E3 Lami1": 12, "E3 Lami2": 13, "E3 Lami3": 14, "E3 Lami4": 15}

    dBookTit = f"{year}{monthSel}.xlsx"
    # Kind of important to fix errors. If i change the template, I dont want to have to modify the file directly
    if os.path.exists(dBookTit):
        os.remove(dBookTit)
    # Creates the template copy. Names it YearMonth.xlsx
    os.system('copy Template.xlsx ' + dBookTit)
    # Open the workbook, initiate the worksheet
    dynamicBook = openpyxl.load_workbook(dBookTit)
    ws1 = dynamicBook["Sheet1"]

    # Archaic and unnecessary                   V
    DLKey = {1: "F", 2: "G", 3: "H", 4: "I", 5: "J", 6: "K", 7: "L", 8: "M", 9: "N", 10: "O",
             11: "P", 12: "Q", 13: "R", 14: "S", 15: "T", 16: "U", 17: "V", 18: "W", 19: "X", 20: "Y",
             21: "Z", 22: "AA", 23: "AB", 24: "AC", 25: "AD", 26: "AE", 27: "AF", 28: "AG", 29: "AH",
             30: "AI", 31: "AJ", 32: "AK", 33: "AL", 34: "AM", 35: "AN", 36: "AO", 37: "AP", 38: "AQ"
             }
    DetailsKey = {"E1 Lami1": "F", "E1 Lami2": "G", "E1 Lami3": "H", "E1 Lami4": "I",
                  "E2 Lami1": "J", "E2 Lami2": "K", "E2 Lami3": "L", "E2 Lami4": "M",
                  "E3 Lami1": "N", "E3 Lami2": "O", "E3 Lami3": "P", "E3 Lami4": "Q"}
    # ancient requirements here could be modified ^

    # First day of the month 0 - sunday, 6 - saturday
    fDay = (calendar.weekday(year, month, 1) + 1) % 7

    # Generate two dictionaries
    abe, deets = loggera(machine)

    # Save information onto main workbook #

    # Names the month in the worksheet If i want to put them together in one file for the year or something
    ws1["F1"] = monthSel

    # IF I WANT TO GET RID OF THE DICTIONARY-KEY HARD CODING, I NEED TO MODIFY THIS CODE VVVV
    if fDay != 0:
        for i in range(fDay):
            ws1[f"{DLKey[i + 1]}2"] = f"CW {date(year, month, 1).isocalendar()[1]}"
    else:
        fDay = 1
    lastpos = "F2"
    for i in abe:
        if date(year, month, i).weekday() % 6 == 0 and date(year, month, i).weekday() != 0:
            if date(year, month, i).isocalendar()[1] == 52:
                ws1[f"{DLKey[i + fDay]}2"] = f"CW {1}"
            else:
                ws1[f"{DLKey[i + fDay]}2"] = f"CW {(date(year, month, int(i)).isocalendar()[1]) + 1}"
        else:
            ws1[f"{DLKey[i + fDay]}{2}"] = f"CW {date(year, month, int(i)).isocalendar()[1]}"
        if type(abe[i]) is dict:
            for j in abe[i]:
                ws1[DLKey[i + fDay] + str(invdictionary2[j])] = int(abe[i][j])
        d = f"{monthSel[0:3]}-{i}"
        ws1[DLKey[i + fDay] + str(3)] = d

    for i in range(fDay + max(sorted(abe))):
        if ws1[lastpos].value != ws1[f"{DLKey[i + 2]}2"].value:
            ws1.merge_cells(f'{lastpos}:{f"{DLKey[i + 1]}2"}')

            lastpos = f"{DLKey[i + 2]}2"

    for i in deets:
        if type(deets[i]) is dict:
            for j in deets[i]:
                indx = i + 17
                string = ""
                for z in deets[i][j]:
                    if len(deets[i][j]) > 1:
                        print(i,j,deets[i][j])

                    string += z + "||"
                ws1[DetailsKey[j] + str(indx)] = string
        d = f"{monthSel[0:3]}-{i}"
        ws1[f"E{i + 17}"] = d
    # IF I WANT TO GET RID OF THE DICTIONARY-KEY HARD CODING, I NEED TO MODIFY THIS CODE   ^^^^^

    # Finish. Save the workbook as YearMonth.xlsx
    dynamicBook.save(dBookTit)


def prerun():
    # If the file doesnt exist in the directory, create file name Year_Month
    if not os.path.exists(genDir):
        os.mkdir(genDir)
    ### Delete everything in the worklog files...
    # We can skip the worklog generation step, actually.
    # Fixed in membrane scraper
    deletewkfiles()

    # This deletes the files in the directory... Seems unnecessary Commented out
    # deleteDirectoryFiles()
    ###
    # If there is nothing in the file Year_Month
    # If we're in the same month and the # of files is less than the current day-1
    # If we are not in the same month and the number of files is less than the # of days in that month
    if not glob2.glob(f'{genDir}/*.xlsx') or \
            (date.today().month == month and len(glob2.glob(f'{genDir}/*.xlsx')) < date.today().day-1) or \
            (date.today().month != month and len(glob2.glob(f'{genDir}/*.xlsx')) < calendar.monthrange(year,month)[1]):
        Demo.grabberFunction(year, month)
        # For now, we will process the workbooks and turn them into the dates afterwards. In the future, I can try a
        # more modular approach. Each day is selected for. If i'm missing worklogs for days 4, 7, and 8,
        # I will put those in a list and download sheets for those specific days.
        WorkBookCompiler()
    # Main function takes the files
    main()


if __name__ == "__main__":
    prerun()
