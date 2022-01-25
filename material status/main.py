import membrane
import time
import openpyxl
import datetime
import glob2
day = datetime.datetime.now().day
month = datetime.datetime.now().month
year = datetime.datetime.now().year
if month < 9:
    month = f"0{month}"
if day < 9:
    day = f"0{day}"
def idworkbook():
    ms = glob2.glob('memstatus/Laminator*.xlsx')
    for i in ms:
        booky = openpyxl.load_workbook(i)
        ws = booky["Laminator Membrane Status"]
        nm = ws["A4"].value
        booky.save(f"memtemp/E{nm[-1]}.xlsx")

def processing():
    MD = {"E1": {"Lami1": {"Teflon": [[], []], "Membrane": []}, "Lami2": {"Teflon": [[], []], "Membrane": []},
                 "Lami3": {"Teflon": [[], []], "Membrane": []}, "Lami4": {"Teflon": [[], []], "Membrane": []}}
        , "E2": {"Lami1": {"Teflon": [[], []], "Membrane": []}, "Lami2": {"Teflon": [[], []], "Membrane": []},
                 "Lami3": {"Teflon": [[], []], "Membrane": []}, "Lami4": {"Teflon": [[], []], "Membrane": []}}
        , "E3": {"Lami1": {"Teflon": [[], []], "Membrane": []}, "Lami2": {"Teflon": [[], []], "Membrane": []},
                 "Lami3": {"Teflon": [[], []], "Membrane": []}, "Lami4": {"Teflon": [[], []], "Membrane": []}}}
    ms = glob2.glob("memtemp/E*")
    for i in ms:
        book = openpyxl.load_workbook(i)
        ws = book["Laminator Membrane Status"]
        for j in list(range(1, 10)):
            # Membrane
            MD[i[8:10]]["Lami1"]["Membrane"] = [ws["E9"].value, ws["E10"].value]
            MD[i[8:10]]["Lami2"]["Membrane"] = [ws["E13"].value, ws["E14"].value]
            MD[i[8:10]]["Lami3"]["Membrane"] = [ws["E17"].value, ws["E18"].value]
            MD[i[8:10]]["Lami4"]["Membrane"] = [ws["E21"].value, ws["E22"].value]
            # Teflons
            MD[i[8:10]]["Lami1"]["Teflon"][0].append(ws[f"{chr(j + 68)}11"].value)
            MD[i[8:10]]["Lami1"]["Teflon"][1].append(ws[f"{chr(j + 68)}12"].value)

            MD[i[8:10]]["Lami2"]["Teflon"][0].append(ws[f"{chr(j + 68)}15"].value)
            MD[i[8:10]]["Lami2"]["Teflon"][1].append(ws[f"{chr(j + 68)}16"].value)

            MD[i[8:10]]["Lami3"]["Teflon"][0].append(ws[f"{chr(j + 68)}19"].value)
            MD[i[8:10]]["Lami3"]["Teflon"][1].append(ws[f"{chr(j + 68)}20"].value)

            MD[i[8:10]]["Lami4"]["Teflon"][0].append(ws[f"{chr(j + 68)}23"].value)
            MD[i[8:10]]["Lami4"]["Teflon"][1].append(ws[f"{chr(j + 68)}24"].value)
    return MD
def listify(a):
    # The goal here is to transform the dictionary into a format that i can use to just place on the excel sheet
    # This might kill me
    E1 = []
    # E1 Lami 1
    E1.extend(a["E1"]["Lami1"]["Membrane"])
    E1.extend(a["E1"]["Lami1"]["Teflon"][0])
    E1.extend(a["E1"]["Lami1"]["Teflon"][1])
    # E1 Lami 2
    E1.extend(a["E1"]["Lami2"]["Membrane"])
    E1.extend(a["E1"]["Lami2"]["Teflon"][0])
    E1.extend(a["E1"]["Lami2"]["Teflon"][1])
    # E1 Lami 3
    E1.extend(a["E1"]["Lami3"]["Membrane"])
    E1.extend(a["E1"]["Lami3"]["Teflon"][0])
    E1.extend(a["E1"]["Lami3"]["Teflon"][1])
    # E1 Lami 4
    E1.extend(a["E1"]["Lami4"]["Membrane"])
    E1.extend(a["E1"]["Lami4"]["Teflon"][0])
    E1.extend(a["E1"]["Lami4"]["Teflon"][1])
    E2 = []
    # E2 Lami 1 ############################
    E2.extend(a["E2"]["Lami1"]["Membrane"])
    E2.extend(a["E2"]["Lami1"]["Teflon"][0])
    E2.extend(a["E2"]["Lami1"]["Teflon"][1])
    # E2 Lami 2
    E2.extend(a["E2"]["Lami2"]["Membrane"])
    E2.extend(a["E2"]["Lami2"]["Teflon"][0])
    E2.extend(a["E2"]["Lami2"]["Teflon"][1])
    # E2 Lami 3
    E2.extend(a["E2"]["Lami3"]["Membrane"])
    E2.extend(a["E2"]["Lami3"]["Teflon"][0])
    E2.extend(a["E2"]["Lami3"]["Teflon"][1])
    # E2 Lami 4
    E2.extend(a["E2"]["Lami4"]["Membrane"])
    E2.extend(a["E2"]["Lami4"]["Teflon"][0])
    E2.extend(a["E2"]["Lami4"]["Teflon"][1])
    E3 = []
    # E3 Lami 1 ##########################
    E3.extend(a["E3"]["Lami1"]["Membrane"])
    E3.extend(a["E3"]["Lami1"]["Teflon"][0])
    E3.extend(a["E3"]["Lami1"]["Teflon"][1])
    # E3 Lami 2
    E3.extend(a["E3"]["Lami2"]["Membrane"])
    E3.extend(a["E3"]["Lami2"]["Teflon"][0])
    E3.extend(a["E3"]["Lami2"]["Teflon"][1])
    # E3 Lami 3
    E3.extend(a["E3"]["Lami3"]["Membrane"])
    E3.extend(a["E3"]["Lami3"]["Teflon"][0])
    E3.extend(a["E3"]["Lami3"]["Teflon"][1])
    # E3 Lami 4
    E3.extend(a["E3"]["Lami4"]["Membrane"])
    E3.extend(a["E3"]["Lami4"]["Teflon"][0])
    E3.extend(a["E3"]["Lami4"]["Teflon"][1])
    omega = []
    omega.extend(E1)
    omega.append(" ")
    omega.extend(E2)
    omega.append(" ")
    omega.extend(E3)
    omega.append(" ")
    return omega

def save(omega):
    wb = openpyxl.load_workbook("Laminator_Consumable Parts_Life Span Trend.xlsx")
    ws = wb["Backdata"]
    i = 2

    while str(ws.cell(row = 34, column = i).value) != f"{year}-{month}-{day} 00:00:00":
        i += 1

    for indx,z in enumerate(omega):
        ws.cell(row = indx+35, column = i).value = z

    wb.save("Laminator_Consumable Parts_Life Span Trend.xlsx")


def main():
    save(listify(processing()))



def run():
    membrane.initialize()
    membrane.download()
    time.sleep(4)
    membrane.driver.quit()

if __name__ == '__main__':
    run()
    idworkbook()
    main()
